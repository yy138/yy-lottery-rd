#!/usr/local/bin/python3
#读取 dlt 开奖数据，进行初步统计
import sys
import xlrd
import xlwt
import string
#import numpy
#import pandas
import openpyxl
#from xlutils.copy import copy
#global r_j

workbook2 = openpyxl.load_workbook('yy-ssq-rd.xlsx') 
sheet4 = workbook2["Cache"]

#读取大乐透开奖数据
workbook = xlrd.open_workbook('ssq-01.xls')
#sheet_names = workbook.sheet_names()
sheet1 = workbook.sheet_by_name('results')
r_numbers = sheet1.nrows

for r_i in range(1,r_numbers):
    rows = sheet1.row_values(r_i)

    #开奖期号
    qihao1 = int(rows[0])
    #qihao2 = 'n'+qihao1
    sheet4.cell(r_i+1,1).value = qihao1

    #前区开奖结果数字
    qian_array1 = rows[1]
    qian_array2 = [0]*6
    for i in range(0,6):
        qian_array2[0+i] = int(qian_array1[0+i*3:0+i*3+2])
    
    for k in range(0,6):
        sheet4.cell(r_i+1,qian_array2[k]+1).value = 1

    #后区开奖结果数字
    hou_array1 = rows[2]
    hou_array2 = [0]
    for j in range(0,1):
        hou_array2[0+j] = int(hou_array1[0+j*3:0+j*3+2])
    
    for l in range(0,1):
        sheet4.cell(r_i+1,hou_array2[l]+1+33).value = 1

    crc = 0
    for r_n in range(2,51):
        count = 1 if type(sheet4.cell(r_i+1,r_n).value) == int else 0
        crc = crc +count
    sheet4.cell(r_i+1,51).value = crc
    print(rows)
'''
    print(qian_array1)
    print(qian_array2)
    print(hou_array1)
    print(hou_array2)
'''
    #写入大乐透开奖数据
    #前区
    #打开原有的表格
   

workbook2.save('yy-ssq-rd.xlsx')






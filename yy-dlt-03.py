#!/usr/local/bin/python3
#读取 dlt 开奖数据，进行初步统计
import sys
import xlrd
import xlwt
import string
#import numpy
#import pandas
import openpyxl
#from xlutils.copy import copy
#global r_j

workbook2 = openpyxl.load_workbook('yy-dlt-rd.xlsx') 
sheet4 = workbook2["Cache"]

#读取大乐透开奖数据
workbook = xlrd.open_workbook('dlt-01.xls')
#sheet_names = workbook.sheet_names()
sheet1 = workbook.sheet_by_name('results')
r_numbers = sheet1.nrows

#大乐透2009015期采用新的开奖机器，增加偏移量设置，
# 如果从第一期2007001开始计算，则 offsize = 0
offsize = 262
offsize2 = offsize -1

for r_i in range(1+offsize2,r_numbers):
    rows = sheet1.row_values(r_i)

    #开奖期号
    qihao1 = int(rows[0])
    #qihao2 = 'n'+qihao1
    sheet4.cell(r_i+1-offsize2,1).value = qihao1

    #前区开奖结果数字
    qian_array1 = rows[1]
    qian_array2 = [0]*5
    for i in range(0,5):
        qian_array2[0+i] = int(qian_array1[0+i*3:0+i*3+2])
    
    for k in range(0,5):
        sheet4.cell(r_i+1-offsize2,qian_array2[k]+1).value = 1

    #后区开奖结果数字
    hou_array1 = rows[2]
    hou_array2 = [0]*2
    for j in range(0,2):
        hou_array2[0+j] = int(hou_array1[0+j*3:0+j*3+2])
    
    for l in range(0,2):
        sheet4.cell(r_i+1-offsize2,hou_array2[l]+1+35).value = 1

    crc = 0
    for r_n in range(2,49):
        count = 1 if type(sheet4.cell(r_i+1-offsize2,r_n).value) == int else 0
        crc = crc +count
    sheet4.cell(r_i+1-offsize2,49).value = crc
    print(rows)
'''
    print(qian_array1)
    print(qian_array2)
    print(hou_array1)
    print(hou_array2)
'''
    #写入大乐透开奖数据
    #前区
    #打开原有的表格
   

workbook2.save('yy-dlt-rd.xlsx')





